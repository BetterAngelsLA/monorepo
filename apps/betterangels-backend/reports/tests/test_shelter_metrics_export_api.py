"""Tests for the shelter metrics export API."""

import datetime
import json
import zipfile
from io import BytesIO

import time_machine
from common.tests.utils import GraphQLBaseTestCase
from django.contrib.auth.models import Permission
from django.urls import reverse
from openpyxl import load_workbook
from organizations.models import Organization
from rest_framework.response import Response
from rest_framework.test import APIClient

from shelters.models import Shelter
from shelters.tests.baker_recipes import shelter_recipe

# Mid-afternoon in LA, the reporting timezone, so "today" is unambiguous.
_FROZEN_NOW = datetime.datetime(2026, 3, 15, 20, 0, tzinfo=datetime.timezone.utc)


class ShelterMetricsExportApiTestCase(GraphQLBaseTestCase):
    def setUp(self) -> None:
        super().setUp()
        self.shelter = shelter_recipe.make(organization=self.org_1)
        self.url = reverse("reports:export_shelter_metrics", kwargs={"shelter_id": str(self.shelter.pk)})
        self._add_shelter_view_permission(self.org_1)
        self.api_client = APIClient()
        self.api_client.force_authenticate(self.org_1_case_manager_1)

    def _add_shelter_view_permission(self, org: Organization) -> None:
        from notes.groups import CASEWORKER

        app_label, codename = Shelter.perms.VIEW.split(".")
        perm = Permission.objects.get(codename=codename, content_type__app_label=app_label)
        org.permission_groups.get(template__name=CASEWORKER.name).group.permissions.add(perm)

    def _get(self, **params: str) -> Response:
        return self.api_client.get(self.url, params, headers={"x-organization-id": str(self.org_1.pk)})

    def test_xlsx_export_names_a_sheet_per_metric(self) -> None:
        response = self._get(export_format="xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            load_workbook(BytesIO(response.content)).sheetnames,
            [
                "daily_occupancy_metrics",
                "daily_bed_status_metrics",
                "reservation_metrics",
                "avg_days_to_occupancy",
            ],
        )

    def test_csv_export_is_a_zip_of_one_file_per_metric(self) -> None:
        response = self._get(export_format="csv")

        self.assertEqual(response["Content-Type"], "application/zip")
        with zipfile.ZipFile(BytesIO(response.content)) as archive:
            self.assertEqual(len(archive.namelist()), 4)

    def test_json_export_returns_titled_blocks(self) -> None:
        response = self._get(export_format="json")

        self.assertEqual(response["Content-Type"], "application/json")
        self.assertEqual(
            [block["title"] for block in json.loads(response.content)],
            [
                "daily_occupancy_metrics",
                "daily_bed_status_metrics",
                "reservation_metrics",
                "avg_days_to_occupancy",
            ],
        )

    def test_response_is_an_attachment_named_for_the_range(self) -> None:
        response = self._get(export_format="json", start_date="2026-06-01", end_date="2026-06-30")

        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="20260601_20260630_shelter_report.json"',
        )

    def test_include_limits_the_export_to_the_named_metrics(self) -> None:
        response = self._get(export_format="xlsx", include="reservation_metrics")

        self.assertEqual(load_workbook(BytesIO(response.content)).sheetnames, ["reservation_metrics"])

    def test_metric_order_does_not_depend_on_the_order_requested(self) -> None:
        """``include`` arrives as a set, so the export pins order to the enum instead."""
        response = self.api_client.get(
            self.url,
            {"export_format": "xlsx", "include": ["reservation_metrics", "daily_occupancy_metrics"]},
            headers={"x-organization-id": str(self.org_1.pk)},
        )

        self.assertEqual(
            load_workbook(BytesIO(response.content)).sheetnames,
            ["daily_occupancy_metrics", "reservation_metrics"],
        )

    @time_machine.travel(_FROZEN_NOW, tick=False)
    def test_omitting_dates_exports_the_last_thirty_days(self) -> None:
        response = self._get(export_format="json")

        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="20260214_20260315_shelter_report.json"',
        )

    def test_rejects_an_unknown_format(self) -> None:
        self.assertEqual(self._get(export_format="pdf").status_code, 400)

    def test_rejects_an_unknown_metric(self) -> None:
        self.assertEqual(self._get(export_format="csv", include="not_a_metric").status_code, 400)

    def test_rejects_a_backwards_date_range(self) -> None:
        response = self._get(export_format="csv", start_date="2026-06-30", end_date="2026-06-01")

        self.assertEqual(response.status_code, 400)

    def test_requires_the_organization_header(self) -> None:
        response = self.api_client.get(self.url, {"export_format": "csv"})

        self.assertEqual(response.status_code, 403)

    def test_requires_authentication(self) -> None:
        self.api_client.force_authenticate(None)

        response = self._get(export_format="csv")

        self.assertIn(response.status_code, (401, 403))

    def test_a_shelter_the_user_cannot_view_is_not_found(self) -> None:
        other_shelter = shelter_recipe.make(organization=self.org_2)
        url = reverse("reports:export_shelter_metrics", kwargs={"shelter_id": str(other_shelter.pk)})

        response = self.api_client.get(url, {"export_format": "csv"}, headers={"x-organization-id": str(self.org_1.pk)})

        self.assertEqual(response.status_code, 404)
