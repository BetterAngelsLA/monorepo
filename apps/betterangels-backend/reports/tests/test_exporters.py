"""Tests for report data export."""

import csv
import zipfile
from datetime import date
from io import BytesIO, StringIO
from typing import Any, cast

import pytest
from model_bakery import baker
from notes.admin import NoteResource
from notes.models import Note
from openpyxl import load_workbook
from shelters.types.reporting import (
    DailyBedStatusMetricsType,
    DailyOccupancyMetricsType,
    ReservationMetricsType,
    ShelterOccupancyMetricsType,
)
from strawberry import ID

from reports.export_options import MetricsExportOptions
from reports.export_to_csv import (
    avg_days_to_occupancy_to_csv,
    csv_files_to_zip,
    daily_bed_status_metrics_to_csv,
    daily_occupancy_metrics_to_csv,
    metrics_to_zip,
    reservation_metrics_to_csv,
    rows_to_csv,
)
from reports.export_to_xlsx import metrics_to_xlsx


def _xlsx_rows(xlsx_content: bytes, worksheet_name: str = "Sheet1") -> list[list[Any]]:
    workbook = load_workbook(BytesIO(xlsx_content))
    worksheet = workbook[worksheet_name]

    return [[value if value is not None else "" for value in row] for row in worksheet.iter_rows(values_only=True)]


def _shelter_occupancy_metrics() -> ShelterOccupancyMetricsType:
    return ShelterOccupancyMetricsType(
        shelter_id=ID("shelter-1"),
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        daily_occupancy=[
            DailyOccupancyMetricsType(
                date=date(2026, 6, 1),
                occupied_count=8,
                total_beds=10,
                occupancy_pct=80.0,
            )
        ],
        daily_bed_status=[
            DailyBedStatusMetricsType(
                date=date(2026, 6, 1),
                available=2,
                occupied=8,
                reserved=1,
                out_of_service=0,
                in_turnaround=0,
            )
        ],
        reservation_metrics=ReservationMetricsType(
            check_in_overdue=3,
            cancelled=2,
            checked_in=11,
            check_in_overdue_to_checked_in=1,
        ),
        avg_days_to_occupancy=4.5,
    )


def _all_metric_export_options() -> list[MetricsExportOptions]:
    return [
        MetricsExportOptions.DAILY_OCCUPANCY_METRICS,
        MetricsExportOptions.DAILY_BED_STATUS_METRICS,
        MetricsExportOptions.RESERVATION_METRICS,
        MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY,
    ]


class TestRowsToCsv:
    def test_exports_rows_with_headers(self) -> None:
        csv_content = rows_to_csv(
            rows=[
                {"date": date(2026, 6, 1), "available": 5},
                {"date": date(2026, 6, 2), "available": 7},
            ],
            headers=["date", "available"],
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["date", "available"],
            ["2026-06-01", "5"],
            ["2026-06-02", "7"],
        ]

    def test_fills_missing_row_values_with_empty_string(self) -> None:
        csv_content = rows_to_csv(
            rows=[
                {"date": date(2026, 6, 1), "available": 5},
                {"date": date(2026, 6, 2)},
            ],
            headers=["date", "available"],
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["date", "available"],
            ["2026-06-01", "5"],
            ["2026-06-02", ""],
        ]

    def test_exports_header_for_empty_rows(self) -> None:
        csv_content = rows_to_csv(rows=[], headers=["date", "available"])

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [["date", "available"]]


class TestShelterMetricsExport:
    def test_daily_occupancy_metrics_to_csv(self) -> None:
        csv_content = daily_occupancy_metrics_to_csv(
            "shelter-1",
            [
                DailyOccupancyMetricsType(
                    date=date(2026, 6, 1),
                    occupied_count=8,
                    total_beds=10,
                    occupancy_pct=80.0,
                ),
                DailyOccupancyMetricsType(
                    date=date(2026, 6, 2),
                    occupied_count=9,
                    total_beds=10,
                    occupancy_pct=90.0,
                ),
            ],
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["date", "shelter_id", "occupied_count", "total_beds", "occupancy_pct"],
            ["2026-06-01", "shelter-1", "8", "10", "80.0"],
            ["2026-06-02", "shelter-1", "9", "10", "90.0"],
        ]

    def test_daily_bed_status_metrics_to_csv(self) -> None:
        csv_content = daily_bed_status_metrics_to_csv(
            "shelter-1",
            [
                DailyBedStatusMetricsType(
                    date=date(2026, 6, 1),
                    available=2,
                    occupied=8,
                    reserved=1,
                    out_of_service=0,
                    in_turnaround=0,
                )
            ],
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["date", "shelter_id", "available", "occupied", "reserved", "out_of_service", "in_turnaround"],
            ["2026-06-01", "shelter-1", "2", "8", "1", "0", "0"],
        ]

    def test_reservation_metrics_to_csv(self) -> None:
        csv_content = reservation_metrics_to_csv(
            "shelter-1",
            date(2026, 6, 1),
            date(2026, 6, 30),
            ReservationMetricsType(
                check_in_overdue=3,
                cancelled=2,
                checked_in=11,
                check_in_overdue_to_checked_in=1,
            ),
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            [
                "start_date",
                "end_date",
                "shelter_id",
                "check_in_overdue",
                "cancelled",
                "checked_in",
                "check_in_overdue_to_checked_in",
            ],
            ["2026-06-01", "2026-06-30", "shelter-1", "3", "2", "11", "1"],
        ]

    def test_avg_days_to_occupancy_to_csv(self) -> None:
        csv_content = avg_days_to_occupancy_to_csv(
            "shelter-1",
            date(2026, 6, 1),
            date(2026, 6, 30),
            4.5,
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["start_date", "end_date", "shelter_id", "avg_days_to_occupancy"],
            ["2026-06-01", "2026-06-30", "shelter-1", "4.5"],
        ]

    def test_avg_days_to_occupancy_to_csv_with_none(self) -> None:
        csv_content = avg_days_to_occupancy_to_csv(
            "shelter-1",
            date(2026, 6, 1),
            date(2026, 6, 30),
            None,
        )

        rows = list(csv.reader(StringIO(csv_content)))

        assert rows == [
            ["start_date", "end_date", "shelter_id", "avg_days_to_occupancy"],
            ["2026-06-01", "2026-06-30", "shelter-1", ""],
        ]

    def test_csv_files_to_zip_includes_each_csv_file(self) -> None:
        zip_content = csv_files_to_zip(
            {
                "daily_occupancy_metrics.csv": "date,total_beds\r\n2026-06-01,10\r\n",
                "reservation_metrics.csv": "start_date,checked_in\r\n2026-06-01,11\r\n",
            }
        )

        with zipfile.ZipFile(BytesIO(zip_content)) as zip_file:
            assert sorted(zip_file.namelist()) == [
                "daily_occupancy_metrics.csv",
                "reservation_metrics.csv",
            ]
            assert zip_file.read("daily_occupancy_metrics.csv").decode() == "date,total_beds\r\n2026-06-01,10\r\n"

    def test_metrics_to_zip_exports_all_metric_files(self) -> None:
        zip_content = metrics_to_zip(
            ShelterOccupancyMetricsType(
                shelter_id=ID("shelter-1"),
                start_date=date(2026, 6, 1),
                end_date=date(2026, 6, 30),
                daily_occupancy=[
                    DailyOccupancyMetricsType(
                        date=date(2026, 6, 1),
                        occupied_count=8,
                        total_beds=10,
                        occupancy_pct=80.0,
                    )
                ],
                daily_bed_status=[
                    DailyBedStatusMetricsType(
                        date=date(2026, 6, 1),
                        available=2,
                        occupied=8,
                        reserved=1,
                        out_of_service=0,
                        in_turnaround=0,
                    )
                ],
                reservation_metrics=ReservationMetricsType(
                    check_in_overdue=3,
                    cancelled=2,
                    checked_in=11,
                    check_in_overdue_to_checked_in=1,
                ),
                avg_days_to_occupancy=4.5,
            ),
            _all_metric_export_options(),
        )

        with zipfile.ZipFile(BytesIO(zip_content)) as zip_file:
            assert sorted(zip_file.namelist()) == [
                "20260601_20260630_avg_days_to_occupancy.csv",
                "20260601_20260630_daily_bed_status_metrics.csv",
                "20260601_20260630_daily_occupancy_metrics.csv",
                "20260601_20260630_reservation_metrics.csv",
            ]
            daily_occupancy_rows = list(
                csv.reader(StringIO(zip_file.read("20260601_20260630_daily_occupancy_metrics.csv").decode()))
            )
            reservation_rows = list(
                csv.reader(StringIO(zip_file.read("20260601_20260630_reservation_metrics.csv").decode()))
            )

        assert daily_occupancy_rows == [
            ["date", "shelter_id", "occupied_count", "total_beds", "occupancy_pct"],
            ["2026-06-01", "shelter-1", "8", "10", "80.0"],
        ]
        assert reservation_rows == [
            [
                "start_date",
                "end_date",
                "shelter_id",
                "check_in_overdue",
                "cancelled",
                "checked_in",
                "check_in_overdue_to_checked_in",
            ],
            ["2026-06-01", "2026-06-30", "shelter-1", "3", "2", "11", "1"],
        ]

    def test_metrics_to_zip_exports_only_selected_metric_files(self) -> None:
        zip_content = metrics_to_zip(
            _shelter_occupancy_metrics(),
            [MetricsExportOptions.DAILY_OCCUPANCY_METRICS, MetricsExportOptions.RESERVATION_METRICS],
        )

        with zipfile.ZipFile(BytesIO(zip_content)) as zip_file:
            assert sorted(zip_file.namelist()) == [
                "20260601_20260630_daily_occupancy_metrics.csv",
                "20260601_20260630_reservation_metrics.csv",
            ]

    def test_metrics_to_zip_rejects_empty_options(self) -> None:
        with pytest.raises(ValueError, match="At least one"):
            metrics_to_zip(_shelter_occupancy_metrics(), [])

    def test_metrics_to_zip_rejects_unknown_options(self) -> None:
        with pytest.raises(ValueError, match="unknown_metric"):
            metrics_to_zip(
                _shelter_occupancy_metrics(),
                [MetricsExportOptions.DAILY_OCCUPANCY_METRICS, cast(MetricsExportOptions, "unknown_metric")],
            )

    def test_metrics_to_xlsx_exports_selected_metrics_as_worksheets(self) -> None:
        filename, xlsx_content = metrics_to_xlsx(
            _shelter_occupancy_metrics(),
            _all_metric_export_options(),
        )

        workbook = load_workbook(BytesIO(xlsx_content))
        assert filename == "20260601_20260630_shelter_report.xlsx"
        assert workbook.sheetnames == [
            "Daily Occupancy",
            "Daily Bed Status",
            "Reservation Metrics",
            "Avg Days To Occupancy",
        ]

        daily_occupancy_rows = _xlsx_rows(xlsx_content, "Daily Occupancy")
        reservation_rows = _xlsx_rows(xlsx_content, "Reservation Metrics")

        assert daily_occupancy_rows == [
            ["date", "shelter_id", "occupied_count", "total_beds", "occupancy_pct"],
            ["2026-06-01", "shelter-1", 8, 10, 80.0],
        ]
        assert reservation_rows == [
            [
                "start_date",
                "end_date",
                "shelter_id",
                "check_in_overdue",
                "cancelled",
                "checked_in",
                "check_in_overdue_to_checked_in",
            ],
            ["2026-06-01", "2026-06-30", "shelter-1", 3, 2, 11, 1],
        ]

    def test_metrics_to_xlsx_exports_only_selected_metric_worksheets(self) -> None:
        filename, xlsx_content = metrics_to_xlsx(
            _shelter_occupancy_metrics(),
            [MetricsExportOptions.DAILY_OCCUPANCY_METRICS, MetricsExportOptions.RESERVATION_METRICS],
        )

        workbook = load_workbook(BytesIO(xlsx_content))
        assert filename == "20260601_20260630_shelter_report.xlsx"
        assert workbook.sheetnames == ["Daily Occupancy", "Reservation Metrics"]

    def test_metrics_to_xlsx_rejects_empty_options(self) -> None:
        with pytest.raises(ValueError, match="At least one"):
            metrics_to_xlsx(_shelter_occupancy_metrics(), [])

    def test_metrics_to_xlsx_rejects_unknown_options(self) -> None:
        with pytest.raises(ValueError, match="unknown_metric"):
            metrics_to_xlsx(
                _shelter_occupancy_metrics(),
                [MetricsExportOptions.DAILY_OCCUPANCY_METRICS, cast(MetricsExportOptions, "unknown_metric")],
            )


class TestNoteResourceExport:
    """Tests for exporting notes via NoteResource (django-import-export)."""

    @pytest.mark.django_db
    def test_export_empty_queryset(self) -> None:
        """Test exporting an empty queryset."""
        resource = NoteResource()
        dataset = resource.export(queryset=Note.objects.none())
        csv_content = dataset.csv

        # Should have headers but no data rows
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 1  # At least headers
        assert "Client ID" in lines[0] or "client_id" in lines[0].lower()

    @pytest.mark.django_db
    def test_export_single_note(self) -> None:
        """Test exporting a single note."""
        note = baker.make(Note)

        resource = NoteResource()
        dataset = resource.export(queryset=Note.objects.filter(pk=note.pk))
        csv_content = dataset.csv

        # Should have headers + 1 data row
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 2

    @pytest.mark.django_db
    def test_export_multiple_notes(self) -> None:
        """Test exporting multiple notes."""
        notes = baker.make(Note, _quantity=5)

        resource = NoteResource()
        dataset = resource.export(queryset=Note.objects.filter(pk__in=[n.pk for n in notes]))
        csv_content = dataset.csv

        # Should have headers + 5 data rows
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 6

    @pytest.mark.django_db
    def test_export_csv_format(self) -> None:
        """Test that exported CSV has correct format."""
        note = baker.make(Note)

        resource = NoteResource()
        dataset = resource.export(queryset=Note.objects.filter(pk=note.pk))
        csv_content = dataset.csv

        # Check that it's valid CSV (has commas)
        assert "," in csv_content

        # Check that it has multiple lines (header + data)
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 2

    def test_note_resource_from_django_import_export(self) -> None:
        """Test that NoteResource is from django-import-export."""
        from import_export.resources import ModelResource

        # NoteResource should be a subclass of ModelResource
        assert issubclass(NoteResource, ModelResource)
