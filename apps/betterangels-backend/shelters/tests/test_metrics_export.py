"""Tests for rendering shelter reporting metrics as downloadable files."""

import json
import zipfile
from datetime import date
from io import BytesIO
from typing import Any, cast

import pytest
from openpyxl import load_workbook
from shelters.types.reporting import (
    DailyBedStatusMetricsType,
    DailyOccupancyMetricsType,
    ReservationMetricsType,
    ShelterOccupancyMetricsType,
)
from strawberry import ID

from shelters.services.metrics_export import MetricsExportOptions, shelter_metrics_export

ALL_OPTIONS = [
    MetricsExportOptions.DAILY_OCCUPANCY_METRICS,
    MetricsExportOptions.DAILY_BED_STATUS_METRICS,
    MetricsExportOptions.RESERVATION_METRICS,
    MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY,
]


def _metrics(avg_days_to_occupancy: float | None = 4.5) -> ShelterOccupancyMetricsType:
    return ShelterOccupancyMetricsType(
        shelter_id=ID("shelter-1"),
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        daily_occupancy=[
            DailyOccupancyMetricsType(date=date(2026, 6, 1), occupied_count=8, total_beds=10, occupancy_pct=80.0)
        ],
        daily_bed_status=[
            DailyBedStatusMetricsType(
                date=date(2026, 6, 1),
                available=2,
                occupied=8,
                reserved=1,
                out_of_service=0,
                in_turnaround=0,
            )
        ],
        reservation_metrics=ReservationMetricsType(
            check_in_overdue=3,
            cancelled=2,
            checked_in=11,
            check_in_overdue_to_checked_in=1,
        ),
        avg_days_to_occupancy=avg_days_to_occupancy,
    )


def _zip_members(body: bytes) -> dict[str, str]:
    with zipfile.ZipFile(BytesIO(body)) as archive:
        return {name: archive.read(name).decode() for name in archive.namelist()}


def _sheet_rows(body: bytes, sheet: str) -> list[list[Any]]:
    workbook = load_workbook(BytesIO(body))
    return [list(row) for row in workbook[sheet].iter_rows(values_only=True)]


class TestFilenameAndContentType:
    def test_csv_is_delivered_as_a_zip_named_for_the_date_range(self) -> None:
        filename, content_type, _ = shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="csv")

        assert filename == "20260601_20260630_shelter_report.zip"
        assert content_type == "application/zip"

    @pytest.mark.parametrize(
        ("export_format", "expected_content_type"),
        [
            ("json", "application/json"),
            ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ],
    )
    def test_single_file_formats_keep_their_extension(self, export_format: str, expected_content_type: str) -> None:
        filename, content_type, _ = shelter_metrics_export(
            metrics=_metrics(), options=ALL_OPTIONS, export_format=export_format
        )

        assert filename == f"20260601_20260630_shelter_report.{export_format}"
        assert content_type == expected_content_type


class TestCsvExport:
    def test_writes_one_file_per_selected_metric(self) -> None:
        _, _, body = shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="csv")

        assert sorted(_zip_members(body)) == [
            "avg_days_to_occupancy.csv",
            "daily_bed_status_metrics.csv",
            "daily_occupancy_metrics.csv",
            "reservation_metrics.csv",
        ]

    def test_daily_occupancy_rows_carry_the_shelter_and_counts(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(),
            options=[MetricsExportOptions.DAILY_OCCUPANCY_METRICS],
            export_format="csv",
        )

        assert _zip_members(body)["daily_occupancy_metrics.csv"].splitlines() == [
            "date,shelter_id,occupied_count,total_beds,occupancy_pct",
            "2026-06-01,shelter-1,8,10,80.0",
        ]

    def test_every_metric_serializes_dates_the_same_way(self) -> None:
        """Occupancy and bed status once disagreed: one wrote isoformat, one a raw date."""
        _, _, body = shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="csv")
        members = _zip_members(body)

        assert members["daily_occupancy_metrics.csv"].splitlines()[1].startswith("2026-06-01,")
        assert members["daily_bed_status_metrics.csv"].splitlines()[1].startswith("2026-06-01,")
        assert members["reservation_metrics.csv"].splitlines()[1].startswith("2026-06-01,2026-06-30,")

    def test_a_missing_average_is_left_blank(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(avg_days_to_occupancy=None),
            options=[MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY],
            export_format="csv",
        )

        assert _zip_members(body)["avg_days_to_occupancy.csv"].splitlines()[1] == "2026-06-01,2026-06-30,shelter-1,"


class TestXlsxExport:
    def test_gives_each_metric_its_own_named_sheet(self) -> None:
        _, _, body = shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="xlsx")

        assert load_workbook(BytesIO(body)).sheetnames == [option.value for option in ALL_OPTIONS]

    def test_bed_status_sheet_holds_the_counts(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(),
            options=[MetricsExportOptions.DAILY_BED_STATUS_METRICS],
            export_format="xlsx",
        )

        header, row = _sheet_rows(body, MetricsExportOptions.DAILY_BED_STATUS_METRICS.value)

        assert header == ["date", "shelter_id", "available", "occupied", "reserved", "out_of_service", "in_turnaround"]
        assert row[1:] == ["shelter-1", 2, 8, 1, 0, 0]


class TestJsonExport:
    def test_emits_one_titled_block_per_metric(self) -> None:
        _, _, body = shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="json")

        assert [block["title"] for block in json.loads(body)] == [option.value for option in ALL_OPTIONS]

    def test_reservation_block_holds_one_row_for_the_range(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(),
            options=[MetricsExportOptions.RESERVATION_METRICS],
            export_format="json",
        )

        assert json.loads(body) == [
            {
                "title": "reservation_metrics",
                "data": [
                    {
                        "start_date": "2026-06-01",
                        "end_date": "2026-06-30",
                        "shelter_id": "shelter-1",
                        "check_in_overdue": 3,
                        "cancelled": 2,
                        "checked_in": 11,
                        "check_in_overdue_to_checked_in": 1,
                    }
                ],
            }
        ]

    def test_a_missing_average_is_null(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(avg_days_to_occupancy=None),
            options=[MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY],
            export_format="json",
        )

        assert json.loads(body)[0]["data"][0]["avg_days_to_occupancy"] is None


class TestOptionSelection:
    def test_exports_only_the_requested_metrics(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(),
            options=[MetricsExportOptions.RESERVATION_METRICS, MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY],
            export_format="xlsx",
        )

        assert load_workbook(BytesIO(body)).sheetnames == ["reservation_metrics", "avg_days_to_occupancy"]

    def test_renders_metrics_in_the_order_requested(self) -> None:
        _, _, body = shelter_metrics_export(
            metrics=_metrics(),
            options=[MetricsExportOptions.AVG_DAYS_TO_OCCUPANCY, MetricsExportOptions.RESERVATION_METRICS],
            export_format="xlsx",
        )

        assert load_workbook(BytesIO(body)).sheetnames == ["avg_days_to_occupancy", "reservation_metrics"]


class TestRejectedInput:
    def test_rejects_an_empty_selection(self) -> None:
        with pytest.raises(ValueError, match="At least one metric export option"):
            shelter_metrics_export(metrics=_metrics(), options=[], export_format="csv")

    def test_rejects_an_unknown_metric(self) -> None:
        with pytest.raises(ValueError, match="Unknown metric export options: unknown_metric"):
            shelter_metrics_export(
                metrics=_metrics(),
                options=[
                    MetricsExportOptions.DAILY_OCCUPANCY_METRICS,
                    cast(MetricsExportOptions, "unknown_metric"),
                ],
                export_format="csv",
            )

    def test_rejects_an_unsupported_format(self) -> None:
        with pytest.raises(ValueError, match="Unsupported export format: pdf"):
            shelter_metrics_export(metrics=_metrics(), options=ALL_OPTIONS, export_format="pdf")
